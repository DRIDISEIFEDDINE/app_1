LAST_DF = None
LAST_EXCEL_EXPORT = None
LAST_EXPORT_FILE = "last_export.xlsx"
import pandas as pd
import numpy as np
from flask import Flask
import hashlib
import re
import traceback
import warnings
import unicodedata
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional, Tuple

_original_md5 = hashlib.md5


def _safe_md5(data=b"", *args, **kwargs):
    kwargs.pop("usedforsecurity", None)
    return _original_md5(data, *args, **kwargs)


hashlib.md5 = _safe_md5


import pandas as pd
import numpy as np
from flask import Flask
from flask import Flask, jsonify, render_template, request, send_file
from geopy.extra.rate_limiter import RateLimiter
from geopy.geocoders import Nominatim
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, String
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from werkzeug.utils import secure_filename

from modules.backlog_technicien_blueprint import bp as backlog_technicien_bp

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
)

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_FOLDER = BASE_DIR / "uploads"
OUTPUT_FOLDER = BASE_DIR / "outputs"
STATIC_FOLDER = BASE_DIR / "static"
TEMPLATE_FOLDER = BASE_DIR / "templates"
ORANGE_LOGO_PATH = BASE_DIR / "Orange1.png"
SITE_ICON_PATH = BASE_DIR / "Orange site.png"
EXCEL_ICON_PATH = BASE_DIR / "Icone Excel.png"

UPLOAD_FOLDER.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)
(STATIC_FOLDER / "img").mkdir(parents=True, exist_ok=True)
TEMPLATE_FOLDER.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {"xlsx", "xls", "xlsm", "xltx", "xltm"}

app = Flask(__name__, template_folder=str(TEMPLATE_FOLDER), static_folder=str(STATIC_FOLDER))
app.register_blueprint(backlog_technicien_bp)
app.config["UPLOAD_FOLDER"] = str(UPLOAD_FOLDER)
app.config["OUTPUT_FOLDER"] = str(OUTPUT_FOLDER)

LAST_EXCEL_EXPORT: Optional[Path] = None

geolocator = Nominatim(user_agent="backlogms_app")
geocode = RateLimiter(geolocator.geocode, min_delay_seconds=1.2)
GEOCODE_CACHE: dict[str, Tuple[Optional[float], Optional[float]]] = {}


def _shared_backlog_loader() -> pd.DataFrame:
    global LAST_EXCEL_EXPORT
    if LAST_EXCEL_EXPORT is None or not LAST_EXCEL_EXPORT.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(LAST_EXCEL_EXPORT, sheet_name="Data", dtype=str, engine="openpyxl")
    except Exception:
        return pd.DataFrame()

app.config["BACKLOGMS_SHARED_LOADER"] = _shared_backlog_loader

TEAM_MAP = {
    "tunis": "Equipe Nord",
    "nabeul": "Equipe Cap-Bon",
    "sousse": "Equipe Centre",
    "ben arous": "Equipe Nord",
    "ariana": "Equipe Nord",
    "manouba": "Equipe Nord",
    "medenine": "Equipe Sud",
    "monastir": "Equipe Centre",
    "jendouba": "Equipe Centre",
    "kef": "Equipe Centre",
    "sfax": "Equipe Sud",
    "bizerte": "Equipe Nord",
    "gabes": "Equipe Sud",
    "zagouan": "Equipe Cap-Bon",
    "kairouan": "Equipe Centre",
    "sidi bouzid": "Equipe Sud",
    "tataouine": "Equipe Sud",
    "tozeur": "Equipe Sud",
    "beja": "Equipe Nord",
    "béja": "Equipe Nord",
    "siliana": "Equipe Centre",
    "kasserine": "Equipe Sud",
    "gafsa": "Equipe Sud",
    "mahdia": "Equipe Sud",
}

GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
BLUE_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
ORANGE_FILL = PatternFill(fill_type="solid", fgColor="F4B183")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="C96B00")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_GREEN_FONT = Font(color="006100", bold=True)
BOLD_BLUE_FONT = Font(color="0000FF", bold=True)
BOLD_RED_FONT = Font(color="9C0006", bold=True)
BOLD_FONT = Font(bold=True)


def allowed_file(filename: str) -> bool:
    if not filename:
        return False

    name = filename.strip()
    if name.startswith("~$"):
        return False

    return "." in name and name.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def now_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def safe_col(df: pd.DataFrame, candidates: list[str], default: str = "") -> str:
    lower_map = {str(c).lower().strip(): c for c in df.columns}
    for cand in candidates:
        key = cand.lower().strip()
        if key in lower_map:
            return lower_map[key]
    return default


def normalize_text(s: str) -> str:
    s = str(s).strip().lower()
    s = s.replace("é", "e").replace("è", "e").replace("ê", "e").replace("à", "a")
    s = re.sub(r"\s+", " ", s)
    return s


def clean_governorate(value) -> str:
    if pd.isna(value):
        return ""
    s = str(value).strip()
    if "\\" in s:
        s = s.split("\\")[-1]
    return s.strip()


def governorate_to_team(governorate: str) -> str:
    if not governorate:
        return ""
    return TEAM_MAP.get(normalize_text(governorate), "")


def clean_product(value) -> str:
    if pd.isna(value):
        return ""

    s = str(value).strip()

    if "_" in s:
        s = s.split("_")[0]

    s = s.strip().upper()

    if "VOIP-GPON" in s:
        return "VOIP-GPON"

    return s

def parse_date_only(value):
    if value is None or pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.to_pydatetime()

    s = str(value).strip()
    if not s or s.lower() in {"nan", "nat", "none"}:
        return None

    s = s.split(";")[0].strip()

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M:%S",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass

    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return None
    return dt.to_pydatetime()


def date_to_str(dt):
    if dt is None or pd.isna(dt):
        return ""
    if isinstance(dt, pd.Timestamp):
        if pd.isna(dt):
            return ""
        return dt.strftime("%Y-%m-%d")
    return dt.strftime("%Y-%m-%d")


def compute_age_in_days(dt):
    if dt is None or pd.isna(dt):
        return None
    if isinstance(dt, pd.Timestamp):
        if pd.isna(dt):
            return None
        dt = dt.to_pydatetime()
    return (datetime.now().date() - dt.date()).days


def smart_clean_address(addr: str) -> str:
    if not addr or str(addr).strip().lower() in {"nan", "none"}:
        return ""

    s = str(addr).strip()
    s = re.sub(r"\s+", " ", s)

    words = s.split()
    cleaned_words = []
    for w in words:
        if not cleaned_words or cleaned_words[-1].lower() != w.lower():
            cleaned_words.append(w)
    s = " ".join(cleaned_words)

    markers = ["Avenue", "Rue", "Boulevard", "Route", "Impasse", "Place", "Av", "Bd"]
    lower_s = s.lower()
    best_pos = None
    for marker in markers:
        pos = lower_s.find(marker.lower())
        if pos != -1 and (best_pos is None or pos < best_pos):
            best_pos = pos

    if best_pos is not None and best_pos > 0:
        prefix = s[:best_pos].strip()
        if len(prefix.split()) >= 2:
            s = s[best_pos:].strip()

    s = re.sub(r"\b(\d+)\s+([A-Za-zÀ-ÿ]+)\s+\1\s+\2\b", r"\1 \2", s, flags=re.IGNORECASE)

    tokens = s.split()
    out = []
    for tok in tokens:
        if not out or out[-1].lower() != tok.lower():
            out.append(tok)
    return " ".join(out).strip()


def geocode_address(address: str) -> Tuple[Optional[float], Optional[float]]:
    if not address:
        return None, None

    if address in GEOCODE_CACHE:
        return GEOCODE_CACHE[address]

    try:
        location = geocode(address)
        if location:
            coords = (location.latitude, location.longitude)
            GEOCODE_CACHE[address] = coords
            return coords
    except Exception:
        pass

    GEOCODE_CACHE[address] = (None, None)
    return None, None


def autosize_worksheet(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 45)


def read_excel_any(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()

    if path.name.startswith("~$"):
        raise ValueError(f"Fichier temporaire ignoré : {path.name}")

    try:
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return pd.read_excel(path, dtype=str, engine="openpyxl")

        if suffix == ".xls":
            return pd.read_excel(path, dtype=str, engine="xlrd")

        raise ValueError(f"Format Excel non supporté : {path.name}")

    except Exception as e:
        raise ValueError(f"Fichier Excel invalide ou corrompu : {path.name}") from e


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    gov_src = safe_col(df, ["Champ complémentaire 10 bis", "Champ complementaire 10 bis"])
    prod_src = safe_col(df, ["Identifiant N° 1", "Identifiant N°1", "Identifiant No 1", "Identifiant N 1"])
    affect_src = safe_col(df, ["Affecté depuis", "Affecte depuis"])
    wf_src = safe_col(df, ["Raison sociale correspondant 1"])
    creation_src = safe_col(df, ["Date de création", "Date création", "Date creation"])
    tech_src = safe_col(df, ["Technicien resp. EDS actif", "Technicien resp EDS actif"])
    addr_src = safe_col(df, ["Adresse", "Adresse client", "Adresse installation"])
    lat_src = safe_col(df, ["Champ complémentaire 4", "Champ complementaire 4"])
    lon_src = safe_col(df, ["Champ complémentaire 5", "Champ complementaire 5"])
    cc3_src = safe_col(df, ["Champ complémentaire 3", "Champ complementaire 3"])

    df["Gouvernorat"] = df[gov_src].apply(clean_governorate) if gov_src else ""
    df["Equipe"] = df["Gouvernorat"].apply(governorate_to_team)
    df["Produit"] = df[prod_src].apply(clean_product) if prod_src else ""
    

    affect_dt = (
        df[affect_src].apply(parse_date_only)
        if affect_src and affect_src in df.columns
        else pd.Series([None] * len(df), index=df.index)
    )
    wf_dt = (
        df[wf_src].apply(parse_date_only)
        if wf_src and wf_src in df.columns
        else pd.Series([None] * len(df), index=df.index)
    )
    creation_dt = (
        df[creation_src].apply(parse_date_only)
        if creation_src and creation_src in df.columns
        else pd.Series([None] * len(df), index=df.index)
    )

    df["Date Affectation"] = affect_dt.apply(date_to_str)
    df["Date WF/TT"] = wf_dt.apply(date_to_str)
    df["Date création normalisée"] = creation_dt.apply(date_to_str)
    df["Age Affectation"] = affect_dt.apply(compute_age_in_days)
    df["Age WF TT"] = [
        compute_age_in_days(wf) if wf is not None and not pd.isna(wf) else compute_age_in_days(cr)
        for wf, cr in zip(wf_dt, creation_dt)
    ]

    if addr_src:
        df["Adresse originale"] = df[addr_src].fillna("").astype(str)
        df["Adresse corrigée"] = df["Adresse originale"].apply(smart_clean_address)
    else:
        df["Adresse originale"] = ""
        df["Adresse corrigée"] = ""

    df["Latitude"] = df[lat_src] if lat_src and lat_src in df.columns else ""
    df["Longitude"] = df[lon_src] if lon_src and lon_src in df.columns else ""
    df["Technicien"] = df[tech_src] if tech_src else ""
    df["Champ complémentaire 3"] = df[cc3_src].fillna("").astype(str) if cc3_src else ""

    return df


def _normalize_header_label(value) -> str:
    s = "" if value is None or pd.isna(value) else str(value)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _read_excel_raw_any(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()

    if path.name.startswith("~$"):
        raise ValueError(f"Fichier temporaire ignoré : {path.name}")

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return pd.read_excel(path, dtype=str, engine="openpyxl", header=None)

    if suffix == ".xls":
        return pd.read_excel(path, dtype=str, engine="xlrd", header=None)

    raise ValueError(f"Format Excel non supporté : {path.name}")


def read_wftt_special_file(path: Path) -> pd.DataFrame:
    raw = _read_excel_raw_any(path)

    if raw.empty:
        raise ValueError("Fichier spécial vide.")

    header_row_idx = None
    max_scan = min(len(raw), 20)

    for i in range(max_scan):
        row_values = [_normalize_header_label(v) for v in raw.iloc[i].tolist()]

        has_phone = any("telephone" in v for v in row_values)
        has_date = any("date reclamation" in v for v in row_values)
        has_ref = any("reference" in v for v in row_values)

        if has_phone and has_date and has_ref:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Entête du fichier spécial WF/TT introuvable.")

    headers = [
        str(v).strip() if v is not None and not pd.isna(v) else f"Colonne_{idx}"
        for idx, v in enumerate(raw.iloc[header_row_idx].tolist())
    ]

    wf = raw.iloc[header_row_idx + 1:].copy()
    wf.columns = headers
    wf = wf.dropna(how="all")
    wf.columns = [str(c).strip() for c in wf.columns]

    rename_map = {}
    for col in wf.columns:
        norm = _normalize_header_label(col)

        if "telephone" in norm:
            rename_map[col] = "N°téléphone"
        elif "reference" in norm:
            rename_map[col] = "Référence"
        elif "date reclamation" in norm:
            rename_map[col] = "Date réclamation"
        elif norm in {"etat central", "etat"}:
            rename_map[col] = "ETAT Central"
        elif norm == "central":
            rename_map[col] = "Central"

    wf = wf.rename(columns=rename_map)

    required = {"N°téléphone", "Référence", "Date réclamation"}
    if not required.issubset(set(wf.columns)):
        raise ValueError("Colonnes minimales du fichier spécial WF/TT non trouvées.")

    wf["N°téléphone"] = wf["N°téléphone"].fillna("").astype(str)
    wf["Référence"] = wf["Référence"].fillna("").astype(str)
    wf["Date réclamation"] = wf["Date réclamation"].fillna("").astype(str)

    if "ETAT Central" not in wf.columns:
        wf["ETAT Central"] = ""

    wf = wf[
        (wf["N°téléphone"].astype(str).str.strip() != "")
        | (wf["Référence"].astype(str).str.strip() != "")
    ].copy()

    return wf


def split_normal_and_wftt_files(file_paths: list[Path]) -> tuple[list[Path], Optional[Path]]:
    normal_files = []
    wf_tt_file = None

    for path in file_paths:
        try:
            wf_preview = read_wftt_special_file(path)

            if (
                wf_tt_file is None
                and not wf_preview.empty
                and {"N°téléphone", "Référence", "Date réclamation"}.issubset(set(wf_preview.columns))
            ):
                wf_tt_file = path
            else:
                normal_files.append(path)

        except Exception:
            normal_files.append(path)

    return normal_files, wf_tt_file


def _normalize_join_key(series: pd.Series) -> pd.Series:
    return (
        series.fillna("")
        .astype(str)
        .str.strip()
        .str.replace(".0", "", regex=False)
        .str.replace("\u00A0", "", regex=False)
        .str.replace(r"[^\d]+", "", regex=True)
    )


def enrich_with_wftt_file(df_merged: pd.DataFrame, wf_tt_path: Optional[Path]) -> pd.DataFrame:
    df = df_merged.copy()

    df["Etat WF TT"] = "Non"
    df["Date Réc"] = ""

    if wf_tt_path is None or df.empty:
        return df

    wf = read_wftt_special_file(wf_tt_path)

    phone_col = safe_col(wf, ["N°téléphone", "N° téléphone", "No téléphone", "Telephone", "Téléphone"])
    date_col = safe_col(wf, ["Date réclamation", "Date reclamation"])

    if not phone_col or not date_col:
        return df

    merged_join_col = safe_col(df, ["Champ complémentaire 3", "Champ complementaire 3"])
    if not merged_join_col:
        return df

    wf["_phone_key"] = _normalize_join_key(wf[phone_col])
    wf["_date_key"] = wf[date_col].fillna("").astype(str).str.strip()

    df["_join_key"] = _normalize_join_key(df[merged_join_col])

    wf = wf[wf["_phone_key"] != ""].copy()
    wf = wf.drop_duplicates(subset=["_phone_key"], keep="first")

    lookup_date = dict(zip(wf["_phone_key"], wf["_date_key"]))

    matched = df["_join_key"].isin(lookup_date.keys())

    df.loc[matched, "Etat WF TT"] = "Retour FSI"
    df.loc[~matched, "Etat WF TT"] = "Non"
    df.loc[matched, "Date Réc"] = df.loc[matched, "_join_key"].map(lookup_date).fillna("")
    df.loc[~matched, "Date Réc"] = ""

    df.drop(columns=["_join_key"], inplace=True, errors="ignore")

    return df


def merge_uploaded_files(file_paths: list[Path]) -> pd.DataFrame:
    normal_files, wf_tt_file = split_normal_and_wftt_files(file_paths)

    frames = []
    for path in normal_files:
        try:
            df = read_excel_any(path)
            df = normalize_dataframe(df)
            df["Source fichier"] = path.name
            frames.append(df)
        except Exception:
            continue

    if not frames:
        return pd.DataFrame()
    
    merged_df = pd.concat(frames, ignore_index=True)
    merged_df = enrich_with_wftt_file(merged_df, wf_tt_file)
    global LAST_DF
    LAST_DF = merged_df
    return merged_df



def enrich_addresses_and_gps(df: pd.DataFrame) -> tuple[pd.DataFrame, int, int]:
    df = df.copy()
    corrected_count = 0
    gps_count = 0

    if "Adresse corrigée" not in df.columns:
        df["Adresse corrigée"] = ""
    if "Latitude" not in df.columns:
        df["Latitude"] = ""
    if "Longitude" not in df.columns:
        df["Longitude"] = ""

    for idx in df.index:
        original = str(df.at[idx, "Adresse originale"]) if "Adresse originale" in df.columns else ""
        corrected = str(df.at[idx, "Adresse corrigée"]) if "Adresse corrigée" in df.columns else ""

        normalized_original = original.strip()
        normalized_corrected = corrected.strip()

        if normalized_corrected and normalized_corrected != normalized_original:
            corrected_count += 1

        lat_val = str(df.at[idx, "Latitude"]).strip()
        lon_val = str(df.at[idx, "Longitude"]).strip()

        need_gps = (
            (not lat_val or lat_val.lower() in {"nan", "none", ""})
            and (not lon_val or lon_val.lower() in {"nan", "none", ""})
        )

        if need_gps:
            lat, lon = geocode_address(corrected or original)
            if lat is not None and lon is not None:
                df.at[idx, "Latitude"] = lat
                df.at[idx, "Longitude"] = lon
                gps_count += 1

    return df, corrected_count, gps_count


def build_technician_product_cards(df: pd.DataFrame) -> list[dict]:

    if "Technicien" not in df.columns or "Produit" not in df.columns:
        return []

    # 🔥 DETECTION INTELLIGENTE COLONNE AGE
    age_col = next(
        (c for c in df.columns if "age" in c.lower()),
        None
    )

    if age_col is None:
        print("❌ Colonne AGE introuvable")
        return []

    print("✅ AGE COLUMN =", age_col)  # DEBUG

    base = df.copy()

    base["Technicien"] = base["Technicien"].fillna("").astype(str).str.strip()
    base["Produit"] = base["Produit"].fillna("").astype(str).str.strip()

    # 🔥 CONVERSION ROBUSTE
    base[age_col] = (
        base[age_col]
        .astype(str)
        .str.replace(",", ".")
        .str.strip()
    )

    base[age_col] = pd.to_numeric(base[age_col], errors="coerce").fillna(0)

    # 🔥 DEBUG CRITIQUE
    print("MAX AGE =", base[age_col].max())
    print("NB >10 =", (base[age_col] > 10).sum())
    print("NB =5 =", (base[age_col] == 5).sum())

    # 🔥 CALCUL CORRECT
    alerts_by_tech = (
        base[base[age_col] > 10]
        .groupby("Technicien")
        .size()
        .to_dict()
    )

    tickets_5j_by_tech = (
        base[base[age_col] == 5]
        .groupby("Technicien")
        .size()
        .to_dict()
    )

    tmp = base[(base["Technicien"] != "") & (base["Produit"] != "")]

    grouped = (
        tmp.groupby(["Technicien", "Produit"])
        .size()
        .reset_index(name="Nombre")
    )

    cards = []

    for tech, g in grouped.groupby("Technicien"):

        cards.append({
            "technicien": tech,
            "alerts10": int(alerts_by_tech.get(tech, 0)),
            "tickets5j": int(tickets_5j_by_tech.get(tech, 0)),
            "details": [
                {
                    "produit": row["Produit"],
                    "nombre": int(row["Nombre"])
                }
                for _, row in g.iterrows()
            ],
        })

    return sorted(cards, key=lambda x: (x["alerts10"], x["tickets5j"]), reverse=True)


def _is_retour_fsi_series(df: pd.DataFrame) -> pd.Series:
    if "Etat WF TT" not in df.columns:
        return pd.Series([False] * len(df), index=df.index)

    normalized = df["Etat WF TT"].fillna("").astype(str).str.strip().str.upper()
    return normalized.isin(["RETOUR FSI", "OUI"])

def dashboard_counts(df: pd.DataFrame) -> dict:

    equipe = (
        df["Equipe"].fillna("").replace("", "Non renseigné").value_counts().sort_values(ascending=False)
        if "Equipe" in df.columns
        else pd.Series(dtype=int)
    )

    gouvernorat = (
        df["Gouvernorat"].fillna("").replace("", "Non renseigné").value_counts().sort_values(ascending=False)
        if "Gouvernorat" in df.columns
        else pd.Series(dtype=int)
    )

    tech = (
        df["Technicien"].fillna("").replace("", "Non renseigné").value_counts().sort_values(ascending=False)
        if "Technicien" in df.columns
        else pd.Series(dtype=int)
    )

    prod = (
        df["Produit"].fillna("").replace("", "Non renseigné").value_counts().sort_values(ascending=False)
        if "Produit" in df.columns
        else pd.Series(dtype=int)
    )

    # =========================
    # ALERTES > 10 jours
    # =========================
    alerts_affect_10 = (
        df[df["Age Affectation"].fillna(0).astype(float) > 10]["Technicien"]
        .fillna("")
        .replace("", "Non renseigné")
        .value_counts()
        .sort_values(ascending=False)
        if "Age Affectation" in df.columns and "Technicien" in df.columns
        else pd.Series(dtype=int)
    )

    # =========================
    # ALERTES WF >= 20
    # =========================
    alerts_wf_20 = (
        df[df["Age WF TT"].fillna(0).astype(float) >= 20]["Technicien"]
        .fillna("")
        .replace("", "Non renseigné")
        .value_counts()
        .sort_values(ascending=False)
        if "Age WF TT" in df.columns and "Technicien" in df.columns
        else pd.Series(dtype=int)
    )

    # =========================
    # FSI
    # =========================
    fsi_by_tech = (
        df[_is_retour_fsi_series(df)]["Technicien"]
        .fillna("")
        .replace("", "Non renseigné")
        .value_counts()
        .sort_values(ascending=False)
        if "Etat WF TT" in df.columns and "Technicien" in df.columns
        else pd.Series(dtype=int)
    )

    # =========================
    # KPI SUPPLEMENTAIRES
    # =========================
    if "Technicien" in df.columns:
        tech_col = df["Technicien"].fillna("").astype(str).str.strip()
        tickets_sans_responsable = int((tech_col == "").sum())
    else:
        tickets_sans_responsable = 0

    # ✅ KPI 5 jours (CORRIGÉ)
    if "Age Affectation" in df.columns:
        age_series = df["Age Affectation"].fillna(0).astype(float)

        tickets_5j = int((df["Age Affectation"].fillna(0).astype(float) == 5).sum())

        tickets_5j_by_tech = (
            df[age_series == 5]["Technicien"]
            .fillna("")
            .replace("", "Non renseigné")
            .value_counts()
            .sort_values(ascending=False)
        )
    else:
        tickets_5j = 0
        tickets_5j_by_tech = pd.Series(dtype=int)

    # =========================
    # RETURN FINAL
    # =========================
    return {
        "equipe": {"labels": equipe.index.tolist(), "values": equipe.values.tolist()},
        "gouvernorat": {"labels": gouvernorat.index.tolist(), "values": gouvernorat.values.tolist()},
        "technicien": {"labels": tech.index.tolist(), "values": tech.values.tolist()},
        "produit": {"labels": prod.index.tolist(), "values": prod.values.tolist()},

        "alertes_affect_10": {
            "labels": alerts_affect_10.index.tolist(),
            "values": alerts_affect_10.values.tolist()
        },

        "alertes_wf_20": {
            "labels": alerts_wf_20.index.tolist(),
            "values": alerts_wf_20.values.tolist()
        },

        "fsi_par_technicien": {
            "labels": fsi_by_tech.index.tolist(),
            "values": fsi_by_tech.values.tolist()
        },

        # ✅ NOUVEAU GRAPHE
        "tickets_5j": {
            "labels": tickets_5j_by_tech.index.tolist(),
            "values": tickets_5j_by_tech.values.tolist()
        },

        "technician_product_cards": build_technician_product_cards(df),

        # =========================
        # KPI GLOBAL
        # =========================
        "kpis": {
            "total_tickets": int(len(df)),

            "tickets_5j": tickets_5j,  # ✅ NOUVEAU KPI

            "total_alerts": int( 
                (df["Age Affectation"].fillna(0).astype(float) > 10).sum()
            ) if "Age Affectation" in df.columns else 0,

            "total_retour_fsi": int(
                _is_retour_fsi_series(df).sum()
            ) if "Etat WF TT" in df.columns else 0,

            "total_wf20": int(
                (df["Age WF TT"].fillna(0).astype(float) >= 20).sum()
            ) if "Age WF TT" in df.columns else 0,

            "tickets_sans_responsable": tickets_sans_responsable,
        },
    }

def _value_only_labels() -> DataLabelList:
    dl = DataLabelList()
    dl.showVal = True
    dl.showPercent = False
    dl.showCatName = False
    dl.showSerName = False
    dl.showLegendKey = False
    return dl


def write_dashboard_sheet(xlsx_path: Path):
    wb = load_workbook(xlsx_path)

    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard")

    ws["A1"] = "Dashboard BacklogMS"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL

    df_tmp = pd.read_excel(xlsx_path, sheet_name="Data", engine="openpyxl")

    def fill_count_table(title: str, col_name: str, start_row: int, start_col: int, filtered_df=None):
        source_df = df_tmp if filtered_df is None else filtered_df
        if col_name not in source_df.columns:
            return None

        series = source_df[col_name].fillna("").replace("", "Non renseigné").value_counts().sort_values(ascending=False)

        ws.cell(row=start_row, column=start_col, value=title).font = BOLD_FONT
        ws.cell(row=start_row + 1, column=start_col, value=col_name).fill = HEADER_FILL
        ws.cell(row=start_row + 1, column=start_col + 1, value="Nombre").fill = HEADER_FILL
        ws.cell(row=start_row + 1, column=start_col).font = WHITE_FONT
        ws.cell(row=start_row + 1, column=start_col + 1).font = WHITE_FONT

        row = start_row + 2
        for label, value in series.items():
            ws.cell(row=row, column=start_col, value=str(label))
            ws.cell(row=row, column=start_col + 1, value=int(value))
            row += 1

        return row - 1

    end0 = fill_count_table(
        "Ticket FSI par Technicien",
        "Technicien",
        3,
        1,
        df_tmp[_is_retour_fsi_series(df_tmp)]
        if "Etat WF TT" in df_tmp.columns
        else df_tmp.iloc[0:0],
    )
    end1 = fill_count_table("Backlog par Equipe", "Equipe", 3, 5)
    end2 = fill_count_table("Tickets par Technicien", "Technicien", 3, 9)
    end3 = fill_count_table("Tickets par Produit", "Produit", 3, 13)
    end4 = fill_count_table(
        "Alertes Affectation > 10 jours",
        "Technicien",
        3,
        17,
        df_tmp[df_tmp["Age Affectation"].fillna(0).astype(float) > 10]
        if "Age Affectation" in df_tmp.columns
        else df_tmp.iloc[0:0],
    )
    end5 = fill_count_table("Tickets par Gouvernorat", "Gouvernorat", 3, 21)
    end6 = fill_count_table(
        "Alertes WF TT >= 20 jours",
        "Technicien",
        3,
        25,
        df_tmp[df_tmp["Age WF TT"].fillna(0).astype(float) >= 20]
        if "Age WF TT" in df_tmp.columns
        else df_tmp.iloc[0:0],
    )

    if end0 and end0 >= 5:
        chart0 = BarChart()
        chart0.type = "bar"
        chart0.style = 10
        chart0.title = "Ticket FSI par Technicien"
        chart0.y_axis.title = "Technicien"
        chart0.x_axis.title = "Nombre"
        data = Reference(ws, min_col=2, min_row=4, max_row=end0)
        cats = Reference(ws, min_col=1, min_row=5, max_row=end0)
        chart0.add_data(data, titles_from_data=True)
        chart0.set_categories(cats)
        chart0.height = 8
        chart0.width = 14
        chart0.dLbls = _value_only_labels()
        ws.add_chart(chart0, "A20")

    if end1 and end1 >= 5:
        chart1 = BarChart()
        chart1.type = "bar"
        chart1.style = 10
        chart1.title = "Backlog par Equipe"
        chart1.y_axis.title = "Equipe"
        chart1.x_axis.title = "Nombre"
        data = Reference(ws, min_col=6, min_row=4, max_row=end1)
        cats = Reference(ws, min_col=5, min_row=5, max_row=end1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.height = 8
        chart1.width = 14
        chart1.dLbls = _value_only_labels()
        ws.add_chart(chart1, "J20")

    if end2 and end2 >= 5:
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "Tickets par Technicien"
        chart2.y_axis.title = "Nombre"
        chart2.x_axis.title = "Technicien"
        data = Reference(ws, min_col=10, min_row=4, max_row=end2)
        cats = Reference(ws, min_col=9, min_row=5, max_row=end2)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.height = 8
        chart2.width = 14
        chart2.dLbls = _value_only_labels()
        ws.add_chart(chart2, "S20")

    if end3 and end3 >= 5:
        chart3 = PieChart()
        chart3.title = "Tickets par Produit"
        labels = Reference(ws, min_col=13, min_row=5, max_row=end3)
        data = Reference(ws, min_col=14, min_row=4, max_row=end3)
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(labels)
        chart3.height = 10
        chart3.width = 12
        chart3.dataLabels = _value_only_labels()
        ws.add_chart(chart3, "AB20")

    if end4 and end4 >= 5:
        chart4 = BarChart()
        chart4.type = "col"
        chart4.style = 12
        chart4.title = "Tickets qui dépassent 10 jours de la date d'affectation"
        chart4.y_axis.title = "Nombre"
        chart4.x_axis.title = "Technicien"
        data = Reference(ws, min_col=18, min_row=4, max_row=end4)
        cats = Reference(ws, min_col=17, min_row=5, max_row=end4)
        chart4.add_data(data, titles_from_data=True)
        chart4.set_categories(cats)
        chart4.height = 8
        chart4.width = 14
        chart4.dLbls = _value_only_labels()
        ws.add_chart(chart4, "A40")

    if end5 and end5 >= 5:
        chart5 = BarChart()
        chart5.type = "col"
        chart5.style = 10
        chart5.title = "Tickets par Gouvernorat"
        chart5.y_axis.title = "Nombre"
        chart5.x_axis.title = "Gouvernorat"
        data = Reference(ws, min_col=22, min_row=4, max_row=end5)
        cats = Reference(ws, min_col=21, min_row=5, max_row=end5)
        chart5.add_data(data, titles_from_data=True)
        chart5.set_categories(cats)
        chart5.height = 8
        chart5.width = 14
        chart5.dLbls = _value_only_labels()
        ws.add_chart(chart5, "J40")

    if end6 and end6 >= 5:
        chart6 = BarChart()
        chart6.type = "col"
        chart6.style = 12
        chart6.title = "Alertes WF TT >= 20 jours"
        chart6.y_axis.title = "Nombre"
        chart6.x_axis.title = "Technicien"
        data = Reference(ws, min_col=26, min_row=4, max_row=end6)
        cats = Reference(ws, min_col=25, min_row=5, max_row=end6)
        chart6.add_data(data, titles_from_data=True)
        chart6.set_categories(cats)
        chart6.height = 8
        chart6.width = 14
        chart6.dLbls = _value_only_labels()
        ws.add_chart(chart6, "S40")

    autosize_worksheet(ws)
    wb.save(xlsx_path)


def _shorten_team_label(label: str) -> str:
    s = str(label).strip()
    mapping = {
        "Equipe Cap-Bon": "Cap-Bon",
        "Equipe Centre": "Centre",
        "Equipe Nord": "Nord",
        "Equipe Sud": "Sud",
    }
    return mapping.get(s, s)


def _compact_axis_label(label: str, max_len: int = 14) -> str:
    s = str(label).strip()
    s = _shorten_team_label(s)

    if len(s) <= max_len:
        return s

    parts = s.split()
    if len(parts) >= 2:
        short = parts[0] + " " + parts[1][0] + "."
        if len(short) <= max_len:
            return short

    return s[:max_len] + "..."


def _build_bar_chart_drawing(
    title: str,
    labels: list[str],
    values: list[int],
    width: int = 760,
    height: int = 360,
    bar_color=colors.HexColor("#F59E0B"),
):
    d = Drawing(width, height)

    d.add(String(20, height - 22, title, fontSize=16, fontName="Helvetica-Bold"))

    if not labels or not values:
        d.add(String(20, height / 2, "Aucune donnée", fontSize=12))
        return d

    clean_labels = [_compact_axis_label(str(v), 14) for v in labels]

    bc = VerticalBarChart()
    bc.x = 70
    bc.y = 85
    bc.height = height - 145
    bc.width = width - 130

    bc.data = [values]
    bc.categoryAxis.categoryNames = clean_labels

    max_val = max(values) if values else 1
    bc.valueAxis.valueMin = 0
    bc.valueAxis.valueMax = max_val + max(1, int(max_val * 0.12))
    bc.valueAxis.valueStep = max(1, int(max_val / 5)) if max_val > 0 else 1

    bc.barLabelFormat = "%d"
    bc.barLabels.nudge = 12
    bc.barLabels.fontSize = 11
    bc.categoryAxis.labels.angle = 30
    bc.categoryAxis.labels.fontSize = 12
    bc.categoryAxis.labels.dy = -12
    bc.valueAxis.labels.fontSize = 12
    bc.bars[0].fillColor = bar_color

    d.add(bc)
    return d


def _build_pie_chart_drawing(
    title: str,
    labels: list[str],
    values: list[int],
    width: int = 760,
    height: int = 420,
):
    d = Drawing(width, height)

    d.add(String(20, height - 22, title, fontSize=16, fontName="Helvetica-Bold"))

    if not labels or not values:
        d.add(String(20, height / 2, "Aucune donnée", fontSize=12))
        return d

    pie = Pie()
    pie.x = 205
    pie.y = 35
    pie.width = 320
    pie.height = 320
    pie.data = values
    pie.labels = [f"{str(lbl)[:18]} ({val})" for lbl, val in zip(labels, values)]
    pie.sideLabels = True
    pie.slices.strokeWidth = 0.6
    pie.slices.fontName = "Helvetica"
    pie.slices.fontSize = 10

    palette = [
        "#0F766E",
        "#7C3AED",
        "#2563EB",
        "#06B6D4",
        "#F59E0B",
        "#EF4444",
        "#84CC16",
        "#EC4899",
        "#8B5CF6",
        "#14B8A6",
    ]
    for i, _ in enumerate(values):
        pie.slices[i].fillColor = colors.HexColor(palette[i % len(palette)])

    pie.slices[0].popout = 6

    d.add(pie)
    return d


def _draw_pdf_header(canvas, doc):
    canvas.saveState()

    page_width, page_height = landscape(A4)

    if ORANGE_LOGO_PATH.exists():
        try:
            canvas.drawImage(
                str(ORANGE_LOGO_PATH),
                doc.leftMargin,
                page_height - 1.9 * cm,
                width=1.5 * cm,
                height=1.5 * cm,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass

    canvas.setFont("Helvetica-Bold", 13)
    canvas.setFillColor(colors.HexColor("#C96B00"))
    canvas.drawString(doc.leftMargin + 1.9 * cm, page_height - 1.15 * cm, "Rapport de Backlog Intervention MS")

    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#666666"))
    canvas.drawRightString(
        page_width - doc.rightMargin,
        page_height - 1.1 * cm,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )

    canvas.setStrokeColor(colors.HexColor("#D9D9D9"))
    canvas.setLineWidth(0.8)
    canvas.line(
        doc.leftMargin,
        page_height - 2.1 * cm,
        page_width - doc.rightMargin,
        page_height - 2.1 * cm,
    )

    canvas.restoreState()


def _build_kpi_table(kpis: dict) -> Table:
    data = [
        [
            "Total tickets",
            "Alertes > 10 jours",
            "Total Ticket Retour FSI",
            "Alertes WF TT >= 20 jours",
            "Tickets sans responsable",
        ],
        [
            str(kpis.get("total_tickets", 0)),
            str(kpis.get("total_alerts", 0)),
            str(kpis.get("total_retour_fsi", 0)),
            str(kpis.get("total_wf20", 0)),
            str(kpis.get("tickets_sans_responsable", 0)),
        ],
    ]

    table = Table(
        data,
        colWidths=[5 * cm, 5 * cm, 5.5 * cm, 5.5 * cm, 5.5 * cm],
        hAlign="CENTER",
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C96B00")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#D6B083")),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#FFF4E8")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    return table


def _build_tickets_without_owner_table(df: pd.DataFrame) -> Table:
    wanted_cols = [
        "Numéro ticket",
        "Champ complémentaire 3",
        "Gouvernorat",
        "Produit",
        "Site client correspondant 1",
        "Age Affectation",
        "Age WF TT",
    ]

    if "Technicien" not in df.columns:
        no_owner_df = df.iloc[0:0].copy()
    else:
        no_owner_df = df[df["Technicien"].fillna("").astype(str).str.strip() == ""].copy()

    available_cols = [c for c in wanted_cols if c in no_owner_df.columns]

    if not available_cols:
        data = [["Aucune donnée disponible"]]
        table = Table(data, colWidths=[26 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF4E8")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D6B083")),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        return table

    display_df = no_owner_df[available_cols].fillna("").astype(str)

    rows = [available_cols]
    for _, row in display_df.iterrows():
        rows.append([str(row[col])[:40] for col in available_cols])

    col_widths = []
    for col in available_cols:
        if col in {"Numéro ticket", "Champ complémentaire 3", "Produit"}:
            col_widths.append(3.0 * cm)
        elif col in {"Age Affectation", "Age WF TT"}:
            col_widths.append(2.2 * cm)
        elif col in {"Site client correspondant 1"}:
            col_widths.append(5.0 * cm)
        else:
            col_widths.append(3.2 * cm)

    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C96B00")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D6B083")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFF9F1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def create_pdf_report(df: pd.DataFrame, path: Path) -> Path:
    dash = dashboard_counts(df)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=2.5 * cm,
        bottomMargin=1.0 * cm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleCenter",
        parent=styles["Title"],
        alignment=TA_CENTER,
        textColor=colors.HexColor("#C96B00"),
        fontName="Helvetica-Bold",
        fontSize=20,
        spaceAfter=14,
    )

    section_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading2"],
        textColor=colors.HexColor("#7C2D12"),
        backColor=colors.HexColor("#E5E7EB"),
        fontName="Helvetica-Bold",
        fontSize=15,
        spaceAfter=10,
        spaceBefore=10,
        leftIndent=6,
        rightIndent=6,
        borderPadding=6,
    )

    normal_style = styles["Normal"]
    normal_style.fontSize = 10

    story = []

    story.append(Paragraph("Rapport de Backlog Intervention MS", title_style))
    story.append(Paragraph(f"Généré le : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph("Cartes KPI", section_style))
    story.append(Spacer(1, 1.0 * cm))
    story.append(_build_kpi_table(dash["kpis"]))
    story.append(Spacer(1, 1.2 * cm))

    story.append(PageBreak())
    story.append(Paragraph("Graphiques principaux", section_style))
    story.append(_build_bar_chart_drawing("Backlog par Equipe", dash["equipe"]["labels"], dash["equipe"]["values"]))
    story.append(Spacer(1, 0.35 * cm))
    story.append(
        _build_bar_chart_drawing(
            "Tickets par Technicien",
            dash["technicien"]["labels"],
            dash["technicien"]["values"],
            bar_color=colors.HexColor("#F59E0B"),
        )
    )

    story.append(PageBreak())
    story.append(Paragraph("Graphiques complémentaires", section_style))
    story.append(_build_pie_chart_drawing("Tickets par Produit", dash["produit"]["labels"], dash["produit"]["values"]))
    story.append(Spacer(1, 0.35 * cm))
    story.append(
        _build_bar_chart_drawing(
            "Alertes WF TT >= 20 jours",
            dash["alertes_wf_20"]["labels"],
            dash["alertes_wf_20"]["values"],
            bar_color=colors.HexColor("#C0392B"),
        )
    )

    story.append(PageBreak())
    story.append(Paragraph("Autres graphiques", section_style))
    story.append(
        _build_bar_chart_drawing(
            "Tickets par Gouvernorat",
            dash["gouvernorat"]["labels"],
            dash["gouvernorat"]["values"],
            bar_color=colors.HexColor("#E28A00"),
        )
    )
    story.append(Spacer(1, 0.35 * cm))
    story.append(
        _build_bar_chart_drawing(
            "Tickets qui dépassent 10 jours de la date d'affectation",
            dash["alertes_affect_10"]["labels"],
            dash["alertes_affect_10"]["values"],
            bar_color=colors.HexColor("#CF3F2C"),
        )
    )

    story.append(PageBreak())
    story.append(Paragraph("Ticket FSI par Technicien", section_style))
    story.append(
        _build_bar_chart_drawing(
            "Ticket FSI par Technicien",
            dash["fsi_par_technicien"]["labels"],
            dash["fsi_par_technicien"]["values"],
            bar_color=colors.HexColor("#8E44AD"),
        )
    )

    story.append(PageBreak())
    story.append(Paragraph("Détails des tickets sans responsable", section_style))
    story.append(_build_tickets_without_owner_table(df))

    doc.build(story, onFirstPage=_draw_pdf_header, onLaterPages=_draw_pdf_header)
    return path


def export_dataframe_to_excel(df: pd.DataFrame, filename: str) -> Path:
    output_path = OUTPUT_FOLDER / filename

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)

    wb = load_workbook(output_path)
    ws = wb["Data"]

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT

    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    col_addr_orig = header_map.get("Adresse originale")
    col_addr_corr = header_map.get("Adresse corrigée")
    col_lat = header_map.get("Latitude")
    col_lon = header_map.get("Longitude")
    col_age_aff = header_map.get("Age Affectation")
    col_age_wf_tt = header_map.get("Age WF TT")

    for row in range(2, ws.max_row + 1):
        if col_addr_orig and col_addr_corr:
            orig = ws.cell(row=row, column=col_addr_orig).value
            corr = ws.cell(row=row, column=col_addr_corr).value
            if corr and str(corr).strip() != str(orig).strip():
                ws.cell(row=row, column=col_addr_corr).fill = GREEN_FILL
                ws.cell(row=row, column=col_addr_corr).font = BOLD_GREEN_FONT

        if col_lat:
            lat_cell = ws.cell(row=row, column=col_lat)
            if lat_cell.value not in (None, ""):
                lat_cell.fill = BLUE_FILL
                lat_cell.font = BOLD_BLUE_FONT

        if col_lon:
            lon_cell = ws.cell(row=row, column=col_lon)
            if lon_cell.value not in (None, ""):
                lon_cell.fill = BLUE_FILL
                lon_cell.font = BOLD_BLUE_FONT

        if col_age_aff:
            try:
                age = float(ws.cell(row=row, column=col_age_aff).value)
                if age > 10:
                    ws.cell(row=row, column=col_age_aff).fill = RED_FILL
                    ws.cell(row=row, column=col_age_aff).font = BOLD_RED_FONT
            except Exception:
                pass

        if col_age_wf_tt:
            try:
                age_wf = float(ws.cell(row=row, column=col_age_wf_tt).value)
                if age_wf >= 20:
                    ws.cell(row=row, column=col_age_wf_tt).fill = ORANGE_FILL
            except Exception:
                pass

    autosize_worksheet(ws)
    wb.save(output_path)

    write_dashboard_sheet(output_path)
    return output_path


def _read_last_export_dataframe():
    global LAST_DF, LAST_EXPORT_FILE

    # priorité mémoire
    if LAST_DF is not None and not LAST_DF.empty:
        return LAST_DF

    # fallback fichier disque
    if os.path.exists(LAST_EXPORT_FILE):
        return pd.read_excel(LAST_EXPORT_FILE)

    raise ValueError("Aucun fichier traité disponible.")


def _coerce_numeric_series(df: pd.DataFrame, col: str) -> pd.Series:
    if col not in df.columns:
        return pd.Series([0] * len(df), index=df.index, dtype=float)
    return pd.to_numeric(df[col], errors="coerce").fillna(0)


def _pick_existing_col(df: pd.DataFrame, candidates: list[str]) -> str:
    return safe_col(df, candidates, "")



def _normalize_etat_wf_for_export(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "Etat WF TT" in df.columns:
        normalized = df["Etat WF TT"].fillna("").astype(str).str.strip().str.upper()
        df.loc[normalized.isin(["OUI", "RETOUR FSI"]), "Etat WF TT"] = "Retour FSI"
    return df

def _build_export_dataframe(df: pd.DataFrame, column_specs: list[tuple[str, list[str]]]) -> pd.DataFrame:
    df = _normalize_etat_wf_for_export(df)
    export_df = pd.DataFrame(index=df.index)

    for output_col, candidates in column_specs:
        source_col = _pick_existing_col(df, candidates)
        if source_col:
            export_df[output_col] = df[source_col]
        else:
            export_df[output_col] = ""

    export_df = export_df.fillna("")
    return export_df


def _export_filtered_excel(df: pd.DataFrame, export_df: pd.DataFrame, prefix: str) -> tuple[Path, str]:
    filename = f"{prefix}_{now_stamp()}.xlsx"
    output_path = OUTPUT_FOLDER / filename

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Détails", index=False)

    wb = load_workbook(output_path)
    ws = wb["Détails"]

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT

    autosize_worksheet(ws)
    wb.save(output_path)

    return output_path, filename


def _detail_ticket_export_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    return _build_export_dataframe(
        df,
        [
            ("Numéro ticket", ["Numéro ticket", "Numero ticket"]),
            ("Champ complémentaire 3", ["Champ complémentaire 3", "Champ complementaire 3"]),
            ("Produit", ["Produit"]),
            ("Date Affectation", ["Date Affectation"]),
            ("Adresse correspondant 1", ["Adresse correspondant 1", "Adresse originale", "Adresse corrigée"]),
            ("Nom correspondant 1", ["Nom correspondant 1"]),
            ("Site client correspondant 1", ["Site client correspondant 1"]),
            ("Gouvernorat", ["Gouvernorat"]),
            ("Equipe", ["Equipe"]),
            ("Age Affectation", ["Age Affectation"]),
            ("Etat WF TT", ["Etat WF TT"]),
            ("Date Réc", ["Date Réc"]),
            ("Mobile correspondant 1", ["Mobile correspondant 1"]),
            ("Date de création", ["Date de création", "Date création", "Date creation", "Date création normalisée"]),
        ],
    )

@app.route("/export_kpi_details/<kind>")
def export_kpi_details(kind):

    try:
        global LAST_DF

        # 🔥 sécuriser kind
        if not kind:
            return jsonify({
                "success": False,
                "message": "Type KPI manquant"
            }), 400

        kind = str(kind).strip().lower()

        # 🔥 récupérer DF
        if LAST_DF is None or LAST_DF.empty:
            return jsonify({
                "success": False,
                "message": "Aucun fichier traité disponible."
            }), 400

        df = LAST_DF.copy()

        if "Age Affectation" in df.columns:
            df["Age Affectation"] = pd.to_numeric(df["Age Affectation"], errors="coerce")
        if "Age WF TT" in df.columns:
            df["Age WF TT"] = pd.to_numeric(df["Age WF TT"], errors="coerce")

        kind = (kind or "").strip().lower()

        if kind == "alerts10":
            filtered = df[_coerce_numeric_series(df, "Age Affectation") > 10].copy()
            export_df = _build_export_dataframe(
                filtered,
                [
                    ("Numéro ticket", ["Numéro ticket", "Numero ticket"]),
                    ("Champ complémentaire 3", ["Champ complémentaire 3", "Champ complementaire 3"]),
                    ("Identifiant N° 1", ["Identifiant N° 1", "Identifiant N°1", "Identifiant No 1", "Identifiant N 1"]),
                    ("Produit", ["Produit"]),
                    ("Gouvernorat", ["Gouvernorat"]),
                    ("Site client correspondant 1", ["Site client correspondant 1"]),
                    ("Nom correspondant 1", ["Nom correspondant 1"]),
                    ("Technicien resp. EDS actif", ["Technicien resp. EDS actif", "Technicien resp EDS actif", "Technicien"]),
                    ("Equipe", ["Equipe"]),
                    ("Age Affectation", ["Age Affectation"]),
                    ("Age WF TT", ["Age WF TT"]),
                    ("Latitude", ["Latitude"]),
                    ("Longitude", ["Longitude"]),
                    ("Etat WF TT", ["Etat WF TT"]),
                    ("Mobile correspondant 1", ["Mobile correspondant 1"]),
                ],
            )
            _, filename = _export_filtered_excel(filtered, export_df, "Age_Affectation_Sup_10_j")
            return send_file(OUTPUT_FOLDER / filename, as_attachment=True, download_name=filename)

        if kind == "retourfsi":
            filtered = df[_is_retour_fsi_series(df)].copy() if "Etat WF TT" in df.columns else df.iloc[0:0].copy()
            export_df = _build_export_dataframe(
                filtered,
                [
                    ("Numéro ticket", ["Numéro ticket", "Numero ticket"]),
                    ("Champ complémentaire 3", ["Champ complémentaire 3", "Champ complementaire 3"]),
                    ("Produit", ["Produit"]),
                    ("Identifiant N° 1", ["Identifiant N° 1", "Identifiant N°1", "Identifiant No 1", "Identifiant N 1"]),
                    ("Mobile correspondant 1", ["Mobile correspondant 1"]),
                    ("Technicien", ["Technicien", "Technicien resp. EDS actif", "Technicien resp EDS actif"]),
                    ("Equipe", ["Equipe"]),
                    ("Gouvernorat", ["Gouvernorat"]),
                    ("Date Réc", ["Date Réc"]),
                    ("Date WF/TT", ["Date WF/TT"]),
                    ("Age WF TT", ["Age WF TT"]),
                    ("Etat WF TT", ["Etat WF TT"]),
                    ("Age Affectation", ["Age Affectation"]),
                ],
            )
            _, filename = _export_filtered_excel(filtered, export_df, "Retour_FSI")
            return send_file(OUTPUT_FOLDER / filename, as_attachment=True, download_name=filename)

        if kind == "wf20":
            filtered = df[_coerce_numeric_series(df, "Age WF TT") >= 20].copy()
            export_df = _build_export_dataframe(
                filtered,
                [
                    ("Numéro ticket", ["Numéro ticket", "Numero ticket"]),
                    ("Champ complémentaire 3", ["Champ complémentaire 3", "Champ complementaire 3"]),
                    ("Produit", ["Produit"]),
                    ("Identifiant N° 1", ["Identifiant N° 1", "Identifiant N°1", "Identifiant No 1", "Identifiant N 1"]),
                    ("Mobile correspondant 1", ["Mobile correspondant 1"]),
                    ("Technicien", ["Technicien", "Technicien resp. EDS actif", "Technicien resp EDS actif"]),
                    ("Equipe", ["Equipe"]),
                    ("Gouvernorat", ["Gouvernorat"]),
                    ("Date Réc", ["Date Réc"]),
                    ("Date WF/TT", ["Date WF/TT"]),
                    ("Age WF TT", ["Age WF TT"]),
                    ("Etat WF TT", ["Etat WF TT"]),
                    ("Age Affectation", ["Age Affectation"]),
                    ("Latitude", ["Latitude"]),
                    ("Longitude", ["Longitude"]),
                ],
            )
            _, filename = _export_filtered_excel(filtered, export_df, "WF_TT_GE_20")
            return send_file(OUTPUT_FOLDER / filename, as_attachment=True, download_name=filename)

        if kind == "sansresp":
            tech_series = df["Technicien"].fillna("").astype(str).str.strip() if "Technicien" in df.columns else pd.Series([""] * len(df), index=df.index)
            filtered = df[tech_series == ""].copy()
            export_df = _build_export_dataframe(
                filtered,
                [
                    ("Numéro ticket", ["Numéro ticket", "Numero ticket"]),
                    ("Champ complémentaire 3", ["Champ complémentaire 3", "Champ complementaire 3"]),
                    ("Produit", ["Produit"]),
                    ("Date Affectation", ["Date Affectation"]),
                    ("Adresse correspondant 1", ["Adresse correspondant 1", "Adresse originale", "Adresse corrigée"]),
                    ("Nom correspondant 1", ["Nom correspondant 1"]),
                    ("Site client correspondant 1", ["Site client correspondant 1"]),
                    ("Gouvernorat", ["Gouvernorat"]),
                    ("Equipe", ["Equipe"]),
                    ("Age Affectation", ["Age Affectation"]),
                    ("Etat WF TT", ["Etat WF TT"]),
                    ("Date Réc", ["Date Réc"]),
                    ("Mobile correspondant 1", ["Mobile correspondant 1"]),
                    ("Date de création", ["Date de création", "Date création", "Date creation", "Date création normalisée"]),
                ],
            )
            _, filename = _export_filtered_excel(filtered, export_df, "Sans_Resp")
            return send_file(OUTPUT_FOLDER / filename, as_attachment=True, download_name=filename)
        if kind == "tickets5j":
            filtered = df[_coerce_numeric_series(df, "Age Affectation") == 5].copy()

            export_df = _build_export_dataframe(
            filtered,
            [
                 ("Numéro ticket", ["Numéro ticket"]),
                 ("Technicien", ["Technicien"]),
                 ("Produit", ["Produit"]),
                 ("Gouvernorat", ["Gouvernorat"]),
                 ("Equipe", ["Equipe"]),
                 ("Age Affectation", ["Age Affectation"]),
            ],
        )

            _, filename = _export_filtered_excel(filtered, export_df, "Tickets_5J")
            return send_file(OUTPUT_FOLDER / filename, as_attachment=True, download_name=filename)
        return jsonify({"success": False, "message": "Type d'export KPI non reconnu."}), 400

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Erreur export détail KPI : {str(e)}"}), 500

@app.route("/export_tech_card_details/<path:tech_name>", methods=["GET"])
def export_tech_card_details(tech_name):
    try:
        df = _read_last_export_dataframe()
        if "Technicien" not in df.columns:
            return jsonify({"success": False, "message": "Colonne Technicien introuvable."}), 400

        requested = (tech_name or "").strip()
        tech_series = df["Technicien"].fillna("").astype(str).str.strip()
        filtered = df[tech_series.str.casefold() == requested.casefold()].copy()

        export_df = _detail_ticket_export_dataframe(filtered)
        safe_name = secure_filename(requested) or "Technicien"
        _, filename = _export_filtered_excel(filtered, export_df, f"Tickets_{safe_name}")
        return send_file(OUTPUT_FOLDER / filename, as_attachment=True, download_name=filename)

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Erreur export détail technicien : {str(e)}"}), 500


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/static_orange_logo")
def static_orange_logo():
    if ORANGE_LOGO_PATH.exists():
        return send_file(ORANGE_LOGO_PATH, mimetype="image/png")
    return "", 204

@app.route("/static_site_icon")
def static_site_icon():
    if SITE_ICON_PATH.exists():
        return send_file(SITE_ICON_PATH, mimetype="image/png")
    if ORANGE_LOGO_PATH.exists():
        return send_file(ORANGE_LOGO_PATH, mimetype="image/png")
    return "", 204


@app.route("/static_excel_icon")
def static_excel_icon():
    if EXCEL_ICON_PATH.exists():
        return send_file(EXCEL_ICON_PATH, mimetype="image/png")
    return "", 204



@app.route("/process", methods=["POST"])
def process_files():
    global LAST_EXCEL_EXPORT, LAST_DF

    try:
        files = request.files.getlist("files")

        if not files:
            return jsonify({"success": False, "message": "Aucun fichier fourni."}), 400

        saved_paths = []

        for file in files:
            if not file:
                continue

            original_name = (file.filename or "").strip()

            if not allowed_file(original_name):
                continue

            filename = secure_filename(original_name) or f"input_{now_stamp()}.xlsx"

            save_path = UPLOAD_FOLDER / f"{datetime.now().strftime('%H%M%S')}_{filename}"

            file.save(save_path)
            saved_paths.append(save_path)

        if not saved_paths:
            return jsonify({"success": False, "message": "Aucun fichier Excel valide."}), 400

        df = merge_uploaded_files(saved_paths)

        if df.empty:
            return jsonify({"success": False, "message": "Aucune donnée exploitable trouvée."}), 400

        # 🔥 IMPORTANT
        LAST_DF = df

        export_name = f"BacklogMS_{now_stamp()}.xlsx"
        export_path = export_dataframe_to_excel(df, export_name)

        LAST_EXCEL_EXPORT = export_path

        # ✅ dashboard calculé une seule fois
        dashboard = dashboard_counts(df)

        return jsonify({
            "success": True,
            "message": "Traitement terminé avec succès.",
            "excel_file": export_name,
            "download_url": f"/download/{export_name}",
            "can_export_pdf": True,
            "dashboard": dashboard,
            "kpis": dashboard.get("kpis", {})
        })

    except Exception as e:
        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "message": f"Erreur lors du traitement : {str(e)}"
        }), 500
@app.route("/correct_addresses_gps", methods=["POST"])
def correct_addresses_gps():
    global LAST_EXCEL_EXPORT

    try:
        if LAST_EXCEL_EXPORT is None or not LAST_EXCEL_EXPORT.exists():
            return jsonify({"success": False, "message": "Aucun fichier traité. Veuillez d'abord importer les données."}), 400

        df = pd.read_excel(LAST_EXCEL_EXPORT, sheet_name="Data", dtype=str, engine="openpyxl")

        if "Age Affectation" in df.columns:
            df["Age Affectation"] = pd.to_numeric(df["Age Affectation"], errors="coerce")
        if "Age WF TT" in df.columns:
            df["Age WF TT"] = pd.to_numeric(df["Age WF TT"], errors="coerce")

        df, corrected_count, gps_count = enrich_addresses_and_gps(df)

        export_name = f"BacklogMS_{now_stamp()}.xlsx"
        export_path = export_dataframe_to_excel(df, export_name)
        LAST_EXCEL_EXPORT = export_path

        return jsonify(
            {
                "success": True,
                "message": f"Correction terminée. Adresses corrigées : {corrected_count}, GPS ajoutés : {gps_count}.",
                "excel_file": export_name,
                "download_url": f"/download/{export_name}",
                "can_export_pdf": True,
                "dashboard": dashboard_counts(df),
                "correction_stats": {
                    "addresses_corrected": corrected_count,
                    "gps_added": gps_count,
                },
            }
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Erreur correction adresse/GPS : {str(e)}"}), 500


@app.route("/export_pdf", methods=["GET"])
def export_pdf():
    try:
        if LAST_EXCEL_EXPORT is None or not LAST_EXCEL_EXPORT.exists():
            return "Erreur : aucun fichier traité disponible pour l'export PDF.", 400

        df = pd.read_excel(LAST_EXCEL_EXPORT, sheet_name="Data", engine="openpyxl")
        pdf_name = f"BacklogMS_Dashboard_{now_stamp()}.pdf"
        pdf_path = OUTPUT_FOLDER / pdf_name
        create_pdf_report(df, pdf_path)

        mode = (request.args.get("mode") or "download").strip().lower()

        if mode in {"inline", "print"}:
            response = send_file(
                pdf_path,
                as_attachment=False,
                mimetype="application/pdf",
            )
        response.headers["Content-Disposition"] = f'inline; filename="{pdf_name}"'

        return response
        response = send_file(
        pdf_path,
        as_attachment=True,
        mimetype="application/pdf"
)
        response.headers["Content-Disposition"] = f'attachment; filename="{pdf_name}"'
        return response
    
    except Exception as e:
        traceback.print_exc()
        return "Erreur lors de l'exportation PDF : " + str(e) + "\n" + traceback.format_exc(), 500

from flask import jsonify
import smtplib
from email.message import EmailMessage

@app.route("/static_mail_icon")
def static_mail_icon():
    return send_file(r"C:\xampp\htdocs\app_1\Icone MAIL.png")
from flask import request, jsonify
from email.message import EmailMessage
from datetime import datetime
import base64
import smtplib
import os

@app.route("/send_mail", methods=["POST"])
def send_mail():
    try:
        data = request.get_json()
        charts = data.get("charts", {})

        global LAST_DF, LAST_EXCEL_EXPORT

        # ================= KPI =================
        if LAST_DF is not None:
            total = len(LAST_DF)
            tickets_5j = len(LAST_DF[LAST_DF["Age Affectation"] == 5])
            alerts10 = len(LAST_DF[LAST_DF["Age Affectation"] > 10])
            wf20 = len(LAST_DF[LAST_DF["Age WF TT"] >= 20])
            retour_fsi = len(LAST_DF[LAST_DF["Etat WF TT"] == "Retour FSI"])
            
        else:
            total = alerts10 = wf20 = retour_fsi = tickets_5j = 0

        now = datetime.now().strftime("%d/%m/%Y %H:%M")

        # ================= EMAIL =================
        msg = EmailMessage()
        msg["Subject"] = f"BacklogMS - {now}"
        msg["From"] = "intervention.orange.tn@gmail.com"
        msg["To"] = "intervention.b2b@orange.com"
        msg["Cc"] = "seifeddine.dridi@orange.com"
        # ================= HTML =================
        html = f"""
<html>
<body style="font-family:Arial; background:#f5f5f5; padding:20px;">
<div style="background:white; padding:20px; border-radius:10px;">

<p><b>Bonjour,</b></p>
<p>Veuillez trouver le backlog actuel.</p>

<h3>📊 KPI Dashboard</h3>

<table style="width:100%; text-align:center; border-spacing:10px;">
<tr>

<td style="background:#fff; padding:15px;">
    <div style="font-size:12px;">Total</div>
    <div style="font-size:20px; color:#f97316;">{total}</div>
</td>

<td style="background:#fff7ed; padding:15px;">
    <div style="font-size:12px;">🔥 = 5 jours</div>
    <div style="font-size:20px; color:#f97316;">{tickets_5j}</div>
</td>

<td style="background:#fee2e2; padding:15px;">
    <div style="font-size:12px;">⚠️ > 10 jours</div>
    <div style="font-size:20px; color:#dc2626;">{alerts10}</div>
</td>

<td style="background:#fee2e2; padding:15px;">
    <div style="font-size:12px;">🚨 WF TT ≥ 20</div>
    <div style="font-size:20px; color:#dc2626;">{wf20}</div>
</td>

<td style="background:#ecfdf5; padding:15px;">
    <div style="font-size:12px;">Retour FSI</div>
    <div style="font-size:20px; color:#16a34a;">{retour_fsi}</div>
</td>

</tr>
</table>

<h3>📈 Graphiques</h3>
"""

        # ================= GRAPHES =================
        titles = {
            "chartTech": "Technicien",
            "chart5Days": "Tickets = 5 jours",
            "chartAlertsAffect10": "Tickets > 10 jours",
            "chartGov": "Gouvernorat",
            "chartProd": "Produit",
        }

        cid_map = {}

        for i, (key, img_data) in enumerate(charts.items()):
            if not img_data or "," not in img_data:
                continue

            cid = f"chart{i}"
            cid_map[key] = cid

            html += f"""
            <p><b>{titles.get(key, key)}</b></p>
            <img src="cid:{cid}" style="width:100%; max-width:600px;">
            """

        html += """
<br><br>
<p>Cordialement</p>

<b>DRIDI Seifeddine</b><br>
Chef de service Intervention Multiservices<br>

</div>
</body>
</html>
"""

        # ================= HTML INJECTION =================
        msg.set_content("BacklogMS")
        msg.add_alternative(html, subtype="html")

        # ================= IMAGES =================
        for key, cid in cid_map.items():
            img_data = charts[key]
            img_bytes = base64.b64decode(img_data.split(",")[1])

            msg.get_body("html").add_related(
                img_bytes,
                maintype="image",
                subtype="png",
                cid=cid
            )

        # ================= PJ EXCEL =================
        print("Excel path:", LAST_EXCEL_EXPORT)

        if LAST_EXCEL_EXPORT and os.path.exists(LAST_EXCEL_EXPORT):
            with open(LAST_EXCEL_EXPORT, "rb") as f:
                msg.add_attachment(
                    f.read(),
                    maintype="application",
                    subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename="Backlog.xlsx"
                )
            print("✅ Excel ajouté")
        else:
            print("❌ Excel NON trouvé")

        # ================= PJ PDF =================
        pdf_path = "backlog_dashboard.pdf"

        if os.path.exists(pdf_path):
            with open(pdf_path, "rb") as f:
                msg.add_attachment(
                    f.read(),
                    maintype="application",
                    subtype="pdf",
                    filename="Dashboard.pdf"
                )
            print("✅ PDF ajouté")
        else:
            print("❌ PDF NON trouvé")

        # ================= SMTP =================
        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
            smtp.starttls()
            smtp.login("intervention.orange.tn@gmail.com", "nckkxuofzbielcdo")
            smtp.send_message(msg)

        return jsonify({"message": "Mail envoyé avec succès"})

    except Exception as e:
        print("ERROR MAIL:", e)
        return jsonify({"error": str(e)}), 500
    
@app.route("/download/<path:filename>")
def download_file(filename):
    full_path = OUTPUT_FOLDER / filename
    if not full_path.exists():
        return "Fichier introuvable.", 404
    return send_file(full_path, as_attachment=True)

@app.route("/upload", methods=["POST"])
def upload():
    global LAST_DF, LAST_EXCEL_EXPORT, LAST_EXPORT_FILE

    try:
        files = request.files.getlist("files")

        if not files:
            return jsonify({"success": False, "message": "Aucun fichier"}), 400

        saved_paths = []

        for file in files:
            if not file or file.filename == "":
                continue

            if not allowed_file(file.filename):
                continue

            filename = secure_filename(file.filename)
            path = UPLOAD_FOLDER / filename
            file.save(path)
            saved_paths.append(path)

        if not saved_paths:
            return jsonify({"success": False, "message": "Fichiers invalides"}), 400

        # 🔥 UTILISER TON PIPELINE EXISTANT
        df = merge_uploaded_files(saved_paths)

        if df.empty:
            return jsonify({"success": False, "message": "Aucune donnée exploitable"}), 400

        # 🔥 sauvegarde mémoire
        LAST_DF = df

        # 🔥 sauvegarde disque
        export_name = f"BacklogMS_{now_stamp()}.xlsx"
        export_path = export_dataframe_to_excel(df, export_name)

        LAST_EXCEL_EXPORT = export_path
        LAST_EXPORT_FILE = str(export_path)

        print("✅ DF sauvegardé :", len(df))

        # 🔥 dashboard complet
        dashboard = dashboard_counts(df)

        return jsonify({
            "success": True,
            "message": "Traitement terminé",
            "rows": len(df),

            # ✅ CRITIQUE POUR TON FRONT
            "dashboard": dashboard,
            "kpis": dashboard.get("kpis", {}),

            # bonus
            "download_url": f"/download/{export_name}"
        })

    except Exception as e:
        print("❌ ERROR UPLOAD:", e)
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

print("DF OK:", LAST_DF is not None)
print("Excel path:", LAST_EXCEL_EXPORT)
print("Exists:", os.path.exists(LAST_EXCEL_EXPORT) if LAST_EXCEL_EXPORT else False)

if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)